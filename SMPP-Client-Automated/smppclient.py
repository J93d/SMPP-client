#!/usr/bin/env python

#Standar libraries called
from time import sleep
import subprocess
import os
import sys
import logging
from logging.handlers import RotatingFileHandler
import logging.config
import select
import openpyxl
import ast

from libraries.bind_receiver import bind_receiver  
from libraries.bind_transmitter import bind_transmitter
from libraries.query_sm import query_sm
from libraries.submit_sm import submit_sm
from libraries.unbind import unbind
from libraries.replace_sm import replace_sm
from libraries.cancel_sm import cancel_sm
from libraries.bind_transceiver import bind_transceiver
from libraries.submit_multi import submit_multi
from libraries.data_sm import data_sm
from libraries.submit_sm_load import submit_sm_load
from libraries.smpp_socket import smpp_socket
from libraries.enquire_link import enquire_link

logging.config.fileConfig('logging_config.ini')

class FunctionNameFilter(logging.Filter):
    def filter(self, record):
        record.funcName = record.funcName if record.funcName else "N/A"
        return True

logger = logging.getLogger(__name__)
for handler in logger.handlers:
    handler.addFilter(FunctionNameFilter())

def get_user_input(timeout=5):
    #Use select to wait for input or timeout
    rlist,_,_=select.select([sys.stdin],[],[],timeout)

    if rlist:
        user_input=int(input())
        return user_input
    else:
        return None

# Function to start and stop trace
def tracer(status,_name=str(''),pid=str(''),filter=str(''),path=str('')):
    #Comment this line for enabling trace. This is currently made only for linux system.
    return None
    if filter is None:
        filter = str('')
    if path is None:
        path = str('')
    if status == 'start':
        command = ['sudo','-s','tcpdump','-i','any',filter,'-w',path+_name+'.pcap']
        try:
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, preexec_fn=os.setsid)
            logger.info(f'Trace started with PID: {os.getpgid(process.pid)}')
            sleep(2)
        except Exception as e:
            logger.error(f'Tcpdump could not be started for the following exception: {e}')
        return str(os.getpgid(process.pid))
    elif status == 'stop':
        try:
            sleep(5)
            subprocess.run(['sudo','-s','pkill','TERM','-g',pid], capture_output=True, text=True, check=True)
            logger.info('Trace stopped.')
        except Exception as e:
            logger.error(f'Exception occured while stopping trace: {e}')
        return None
    else:
        return None
        
if __name__=="__main__":
    n=0

    print('Welcome to SMPP Client!\n')
    try:
        host=input('Enter the IP address: ')
        port=input('Enter the port: ')
    except KeyboardInterrupt:
        logger.error("KeyboardInterrupt caught. exiting gracefully.")
        sys.exit(1)

    address=(host,int(port))

    try:
        status_sock_conn=smpp_socket.conn(address)
    except Exception as e:
        logger.error("Socket connection failed with exception: {}".format(e))
        sys.exit(1)

    if status_sock_conn == True:
        logger.info("SMPP Connection successfully created")
    else:
        logger.error("SMPP Connection failed with exception: {}".format(status_sock_conn))
        sys.exit(1)


    try:
        while n > 0:
            try:
                workbook = openpyxl.load_workbook('Testcases.xlsx')
                sheet_account = workbook["Account"]
                sheet_submit = workbook["SubmitSM"]
                sheet_trace = workbook["Information"]
            except Exception as e:
                logger.error(f"Opening Workbook failed with exception: {e}")

            # Defining the inputs to avoid issue if they are missing
            trace_filter = str('')
            trace_path = str('')
            
            for cell in sheet_trace.iter_rows(min_row=2, max_col=2, values_only=True):
                if cell[0] == "Trace Filter":
                    trace_filter = cell[1]
                elif cell[0] == "Path":
                    trace_path = cell[1]
                else:
                    continue

            logger.info(f'Trace filter used: {trace_filter} in path {trace_path}')
            
            # Bind
            headers_account = [cell.value for cell in sheet_account[1][1:]]
            
            for row_index, cell in enumerate(sheet_account.iter_rows(min_row=2, min_col=1, values_only=True), start=2):
                print(f"{row_index-1}: {cell[0]}")
            print("\nEnter the Bind Type: ",end="")

            user_input_row = get_user_input(timeout=10)

            if str("receiver") in sheet_account[f"A{user_input_row+1}"].value.lower():
                bind_receiver(system_id=sheet_account[f"B{user_input_row+1}"].value,password=sheet_account[f"C{user_input_row+1}"].value)
            elif str("transmitter") in sheet_account[f"A{user_input_row+1}"].value.lower():
                bind_transmitter(system_id=sheet_account[f"B{user_input_row+1}"].value,password=sheet_account[f"C{user_input_row+1}"].value)
            elif str("transceiver") in sheet_account[f"A{user_input_row+1}"].value.lower():
                bind_transceiver(system_id=sheet_account[f"B{user_input_row+1}"].value,password=sheet_account[f"C{user_input_row+1}"].value)

            try:
                while n<=0:
                    headers_submit = [cell.value for cell in sheet_submit[1][1:]]
                    print("Test Cases:")
                    for row_index, cell in enumerate(sheet_submit.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
                        print(f"{row_index-1}: {cell[0]}")
                    print("{sheet_submit.max_row}: QuerySM\n{sheet_submit.max_row+1}: CancelSM\n{sheet_submit.max_row+2}: ReplcaSM\n{sheet_submit.max_row+3}: Unbind\n{sheet_submit.max_row+4}: Run All\nEnter the Test Case to be executed: "),end="")

                    user_input_row = get_user_input()

                    while user_input_row is None:
                        try:
                            enquire_link()
                        except Exception as e:
                            logger.error(f"Exception occurred: {e}")
                        user_input_row = get_user_input()

                    logger.info(f"{user_input_row} input received")

                    if int(user_input_row)==(sheet_submit.max_row):
                        process = tracer(status='start', filter=trace_filter,_name='QuerySM')
                        query_sm()
                        tracer(status='stop', pid=process)
                    elif int(user_input_row)==(sheet_submit.max_row+1):
                        process = tracer(status='start', filter=trace_filter,_name='CancelSM')
                        cancel_sm()
                        tracer(status='stop', pid=process)
                    elif int(user_input_row)==(sheet_submit.max_row+2):
                        process = tracer(status='start', filter=trace_filter,_name='ReplaceSM')
                        replace_sm()
                        tracer(status='stop', pid=process)
                    elif int(user_input_row)==(sheet_submit.max_row+3):
                        unbind()
                        sleep(2)
                        smpp_socket.disconnect(close=1)
                        workbook.close()
                        n+=1
                    elif int(user_input_row)==(sheet_submit.max_row+4):
                        test_case_list = [cell[0].value for cell in sheet_submit[f"A{2}:A{100}"] if cell[0].value is not None]

                        for tc in test_case_list:
                            # Start Trace
                            process = tracer(status='start', filter=trace_filter,_name=tc)
                            test_case = test_case_list.index(tc)+2
                            row_value = [cell.value for cell in sheet_submit[f"A{test_case}:T{test_case}"][0]]
                            formatted_values = ', '.join([f'{header}="{value}"' for header, value in zip(headers_submit, row_values[1:]) if value is not None])
                            ast.literal_eval(f"submit_sm({formatted_values})")
                            logger.error(f"{row_values[0]} Test Case Executed.")
                            tracer(status='stop',pid=process)
                        unbind()
                        sleep(2)
                        smpp_socket.disconnect(close=1)
                        workbook.close()
                        n+=1
                    elif int(user_input_row)<(sheet_submit.max_row):
                        process = tracer(status='start',filter=trace_filter,_name=sheet_submit[f"A{user_input_row}"][0]]
                        formatted_values = ', '.join([f'{header}="{value}"' for header, value in zip(headers_submit, row_values[1:]) if value is not None])
                        ast.literal_eval(f"submit_sm({formatted_values})")
                        logger.info(f"{row_values[0]} Test Case Executed.")
                        tracer(status='stop',pid=process)
                    else:
                        logger.error("Wrong Input. Please try again.")            
    except KeyboardInterrupt:
        try:
            smpp_socket.disconnect(close=1)
            sys.exit(1)
        except Exception as e:
            logger.warning("Exception in socket disconnection: {}".format(e))
    except Exception as e:
        logger.error(f"Exception raised: {e}")

